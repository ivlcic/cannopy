import importlib
import os
import subprocess
import sys
from datetime import datetime
from typing import List, Dict, Optional, Any

from app.common import PathLike
from app.pip import Pip


class ClusterExcel:

    def __init__(self, file: PathLike):
        self.file = file
        Pip.install_packages('openpyxl', '3.1.5')
        # noinspection PyUnresolvedReferences
        if importlib.util.find_spec(pkg) is None:
            subprocess.check_call([sys.executable, "-m", "pip", "install", f'{pkg}=={ver}'])

        # noinspection PyPackageRequirements,PyUnresolvedReferences
        from openpyxl import Workbook
        # noinspection PyPackageRequirements,PyUnresolvedReferences
        from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill

        self.wb = Workbook(write_only=True)

        arial = Font(name='Arial', size=10)
        default_style = NamedStyle(name='def')
        default_style.font = arial
        self.wb.add_named_style(default_style)

        default_style_b = NamedStyle(name='def_b')
        default_style_b.font = Font(name='Arial', size=10, b=True)
        default_style_b.fill = PatternFill('solid', fgColor='FEE135')
        self.wb.add_named_style(default_style_b)

        link_style = NamedStyle(name='hl')
        link_style.font = Font(name='Arial', size=10, color='0000FF', underline='single')
        self.wb.add_named_style(link_style)

        link_style_b = NamedStyle(name='hl_b')
        link_style_b.font = Font(name='Arial', size=10, color='0000FF', underline='single', b=True)
        link_style_b.fill = PatternFill('solid', fgColor='FEE135')
        self.wb.add_named_style(link_style_b)

        uuid_style = NamedStyle(name='uuid')
        uuid_style.alignment = Alignment(horizontal='center')
        uuid_style.font = Font(name='Courier New', size=10)
        self.wb.add_named_style(uuid_style)

        published_style = NamedStyle(name='pub')
        published_style.number_format = 'dd. mm. yyyy'
        published_style.font = arial
        published_style.alignment = Alignment(horizontal='center')
        self.wb.add_named_style(published_style)

        broadcast_style = NamedStyle(name='br')
        broadcast_style.number_format = 'HH:mm'
        broadcast_style.font = arial
        broadcast_style.alignment = Alignment(horizontal='center')
        self.wb.add_named_style(broadcast_style)

        created_style = NamedStyle(name='cr')
        created_style.number_format = 'dd. mm. yyyy HH:mm:ss'
        created_style.font = arial
        created_style.alignment = Alignment(horizontal='center')
        self.wb.add_named_style(created_style)

    def cluster_print_sheet(self, sheet_name: Optional[str], cluster_bucket: List[Dict[str, Any]]):
        # noinspection PyPackageRequirements,PyUnresolvedReferences
        from openpyxl.cell import Cell, WriteOnlyCell
        # noinspection PyPackageRequirements,PyUnresolvedReferences
        from openpyxl.styles import Border, Side

        if sheet_name:
            ws = self.wb.create_sheet(sheet_name)
        else:
            ws = self.wb.create_sheet()

        def _xlsx_cluster_cell_border(c_size: int, x: int, x_cell: Cell):
            if x == 0 and c_size > 1:
                # noinspection PyDunderSlots, PyUnresolvedReferences
                x_cell.border = Border(
                    top=Side(style='thin'),
                )

        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 6
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 38

        ws.append([
            # A        B        C            D            E       F         G          H
            'Cluster', 'Title', 'Published', 'Broadcast', 'Type', 'Source', 'Created', 'UUID'
        ])
        # Freeze the first row
        ws.freeze_panes = 'A2'

        for cluster in cluster_bucket:
            articles: List[Dict[str, Any]] = cluster['articles']
            size = len(articles)
            for x, article in enumerate(articles):
                created = datetime.fromisoformat(article['created'].replace('Z', '+00:00')).astimezone()
                published = datetime.fromisoformat(article['published'].replace('Z', '+00:00')).astimezone()

                broadcast = ''
                if article['type'] == 'tv' or article['type'] == 'radio':
                    broadcast = published.replace(tzinfo=None)

                kl_token = os.environ.get('KMAP_TOKEN', None)
                kl_url = os.environ.get('KMAP_URL', None)
                preview_url = ''
                pdf_url = ''
                url = ''
                if kl_token:
                    rel_path = os.path.join(
                        str(created.year), f"{created.month:02d}", f"{created.day:02d}", article['uuid']
                    )
                    params = '&purpose=2&language=en&summaryType=override&showHighlights=true'
                    if article['url']:
                        url = f'{kl_url}/redirect?filePath={rel_path}&{params}&dcStringToken={kl_token}'

                    preview_url = f'{kl_url}/html?filePath={rel_path}&{params}&dcStringToken={kl_token}'
                    pdf_url = f'{kl_url}/pdf?filePath={rel_path}&{params}&dcStringToken={kl_token}'

                row = []
                a_cell = WriteOnlyCell(ws, value=cluster['id'])
                a_cell.style = 'uuid'
                _xlsx_cluster_cell_border(size, x, a_cell)
                row.append(a_cell)

                a_cell = WriteOnlyCell(ws, value=article['title'])
                if article['id'] == cluster['id'] and size > 1:
                    a_cell.style = 'hl_b'
                else:
                    a_cell.style = 'hl'
                a_cell.hyperlink = preview_url

                _xlsx_cluster_cell_border(size, x, a_cell)
                row.append(a_cell)

                a_cell = WriteOnlyCell(ws, value=published.replace(tzinfo=None))
                a_cell.style = 'pub'
                _xlsx_cluster_cell_border(size, x, a_cell)
                row.append(a_cell)

                a_cell = WriteOnlyCell(ws, value=broadcast)
                a_cell.style = 'br'
                _xlsx_cluster_cell_border(size, x, a_cell)
                row.append(a_cell)

                a_cell = WriteOnlyCell(ws, value=article['type'])
                a_cell.style = 'def'
                _xlsx_cluster_cell_border(size, x, a_cell)
                row.append(a_cell)

                a_cell = WriteOnlyCell(ws, value=article['source'])
                a_cell.style = 'def'
                if url:
                    a_cell.style = 'hl'
                    a_cell.hyperlink = url
                else:
                    a_cell.style = 'def'
                _xlsx_cluster_cell_border(size, x, a_cell)
                row.append(a_cell)

                a_cell = WriteOnlyCell(ws, value=created.replace(tzinfo=None))
                a_cell.style = 'cr'
                _xlsx_cluster_cell_border(size, x, a_cell)
                row.append(a_cell)

                a_cell = WriteOnlyCell(ws, value=article['uuid'])
                a_cell.style = 'uuid'
                a_cell.hyperlink = pdf_url
                _xlsx_cluster_cell_border(size, x, a_cell)
                row.append(a_cell)
                ws.append(row)

    def write_xlsx(self, cluster_list: List[Dict[str, Any]]):
        for cluster_bucket in cluster_list:
            self.cluster_print_sheet(cluster_bucket['key'], cluster_bucket['clusters'])
        self.wb.save(self.file)
